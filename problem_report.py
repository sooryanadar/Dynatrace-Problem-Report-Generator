import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkcalendar import DateEntry
import pandas as pd
import requests
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def convert_to_timestamp(date, time_str):
    datetime_str = f"{date} {time_str}"
    return int(time.mktime(pd.to_datetime(datetime_str).timetuple()) * 1000)


def fetch_data():
    try:
        base_url = url_entry.get()
        if '?' not in base_url:
            base_url += '?'
        else:
            base_url += '&'
        base_url += 'pageSize=500'

        from_date = from_date_picker.get_date().strftime("%Y-%m-%d")
        from_time = f"{from_hour_combobox.get()}:{from_minute_combobox.get()}"
        to_date = to_date_picker.get_date().strftime("%Y-%m-%d")
        to_time = f"{to_hour_combobox.get()}:{to_minute_combobox.get()}"
        management_zone = management_zone_entry.get()
        token = token_entry.get()

        if not base_url or not management_zone or not token:
            messagebox.showerror("Error", "All fields must be filled.")
            return

        from_time_timestamp = convert_to_timestamp(from_date, from_time)
        to_time_timestamp = convert_to_timestamp(to_date, to_time)

        params = {
            "from": from_time_timestamp,
            "to": to_time_timestamp,
            "problemSelector": f'managementZones("{management_zone}")'
        }
        headers = {
            "Authorization": f"Api-Token {token}"
        }

        all_problems = []
        page = 1

        while True:
            params["page"] = page
            response = requests.get(base_url, params=params, headers=headers, verify=False)

            if response.status_code == 200:
                data = response.json()
                all_problems.extend(data.get("problems", []))
                total_count = data.get("totalCount", len(all_problems))

                if len(all_problems) >= total_count:
                    break

                page += 1
            else:
                messagebox.showerror("Error", f"Failed to fetch data. Status code: {response.status_code}, Response: {response.text}")
                return

        problems = [
            {
                "Problem ID": problem.get("problemId"),
                "Display ID": problem.get("displayId"),
                "Title": problem.get("title"),
                "Impact Level": problem.get("impactLevel"),
                "Severity Level": problem.get("severityLevel"),
                "Status": problem.get("status"),
                "Root Cause Entity": problem.get("rootCauseEntity", {}).get("name") if problem.get("rootCauseEntity") else "N/A",
                "Start Time": pd.to_datetime(problem.get("startTime"), unit='ms'),
                "End Time": pd.to_datetime(problem.get("endTime"), unit='ms') if problem.get("endTime") != -1 else "Ongoing",
                "Management Zones": ", ".join(zone.get("name", "N/A") for zone in problem.get("managementZones", []))
            }
            for problem in all_problems
        ]

        df = pd.DataFrame(problems)

        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_file:
            return

        wb = Workbook()
        ws_raw = wb.active
        ws_raw.title = "Raw Data"

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        ws_raw.auto_filter.ref = ws_raw.dimensions

        ws_pivot = wb.create_sheet(title="Repetitive Index")
        pivot_data = df.groupby(["Impact Level", "Severity Level"]).size().reset_index(name="Count")

        for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_pivot.cell(row=r_idx, column=c_idx, value=value)

        tab = Table(displayName="PivotTable1", ref=f"A1:C{len(pivot_data) + 1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws_pivot.add_table(tab)

        wb.save(output_file)
        messagebox.showinfo("Success", f"Data saved to {output_file}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")



root = tk.Tk()
root.title("Dynatrace Data Fetcher")
root.geometry("600x500")

style = ttk.Style()
style.configure("TLabel", font=("Arial", 10), padding=5)
style.configure("TButton", font=("Arial", 12), padding=5)
style.configure("TEntry", padding=5)

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

ttk.Label(frame, text="Dynatrace API URL:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
url_entry = ttk.Entry(frame, width=50)
url_entry.grid(row=0, column=1, pady=5, sticky=tk.W)

ttk.Label(frame, text="For Saas (e.g., https://<your-environment-id>.live.dynatrace.com/api/v2/problems)", font=("Arial", 8)).grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
ttk.Label(frame, text="For Managed (e.g., https://<your-domain-name>/e/{your-environment-id}/api/v2/problems)", font=("Arial", 8)).grid(row=2, column=1, sticky=tk.W, pady=(0, 10))

ttk.Label(frame, text="From Date:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
from_date_picker = DateEntry(frame, width=20, date_pattern="yyyy-mm-dd")
from_date_picker.grid(row=3, column=1, pady=5, sticky=tk.W)

ttk.Label(frame, text="From Time:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
from_hour_combobox = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=5)
from_hour_combobox.grid(row=4, column=1, sticky=tk.W, padx=(0, 5))
from_hour_combobox.set("00")
from_minute_combobox = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(60)], width=5)
from_minute_combobox.grid(row=4, column=1, padx=(60, 0), sticky=tk.W)
from_minute_combobox.set("00")

ttk.Label(frame, text="To Date:").grid(row=5, column=0, sticky=tk.W, padx=5, pady=5)
to_date_picker = DateEntry(frame, width=20, date_pattern="yyyy-mm-dd")
to_date_picker.grid(row=5, column=1, pady=5, sticky=tk.W)

ttk.Label(frame, text="To Time:").grid(row=6, column=0, sticky=tk.W, padx=5, pady=5)
to_hour_combobox = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=5)
to_hour_combobox.grid(row=6, column=1, sticky=tk.W, padx=(0, 5))
to_hour_combobox.set("23")
to_minute_combobox = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(60)], width=5)
to_minute_combobox.grid(row=6, column=1, padx=(60, 0), sticky=tk.W)
to_minute_combobox.set("59")

ttk.Label(frame, text="Management Zone:").grid(row=7, column=0, sticky=tk.W, padx=5, pady=5)
management_zone_entry = ttk.Entry(frame, width=50)
management_zone_entry.grid(row=7, column=1, pady=5, sticky=tk.W)
ttk.Label(frame, text="e.g., Production, Staging", font=("Arial", 8)).grid(row=8, column=1, sticky=tk.W)

ttk.Label(frame, text="API Token:").grid(row=9, column=0, sticky=tk.W, padx=5, pady=5)
token_entry = ttk.Entry(frame, width=50, show="*")
token_entry.grid(row=9, column=1, pady=5, sticky=tk.W)
ttk.Label(frame, text="Enter your API Token here", font=("Arial", 8)).grid(row=10, column=1, sticky=tk.W)


ttk.Button(frame, text="Fetch Data", command=fetch_data).grid(row=11, column=0, columnspan=2, pady=20)

author_label = tk.Label(root, text="Author: Soorya Muthuraj Nadar", font=("Arial", 4))
author_label.grid(row=100, column=0, columnspan=2, pady=5, sticky=tk.W)

root.mainloop()
