from flask import Flask, render_template, request, send_file
import pandas as pd
import requests
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)

def convert_to_timestamp(date_str):
    
    return int(time.mktime(pd.to_datetime(date_str).timetuple()) * 1000)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/fetch', methods=['POST'])
def fetch_data():
    try:
        
        url = request.form['url']
        from_time = request.form['from']
        to_time = request.form['to']
        management_zone = request.form['managementZone']
        token = request.form['token']

        
        from_time_timestamp = convert_to_timestamp(from_time)
        to_time_timestamp = convert_to_timestamp(to_time)

        
        params = {
            "from": from_time_timestamp,
            "to": to_time_timestamp,
            "problemSelector": f'managementZones("{management_zone}")'
        }
        headers = {
            "Authorization": f"Api-Token {token}"
        }

        # Fetch data from the API
        response = requests.get(url, params=params, headers=headers)

        if response.status_code == 200:
            data = response.json()
            problems = []

            
            for problem in data.get("problems", []):
                problem_info = {
                    "Problem ID": problem.get("problemId"),
                    "Display ID": problem.get("displayId"),
                    "Title": problem.get("title"),
                    "Impact Level": problem.get("impactLevel"),
                    "Severity Level": problem.get("severityLevel"),
                    "Status": problem.get("status"),
                    "Root Cause Entity": problem.get("rootCauseEntity", {}).get("name") if problem.get("rootCauseEntity") else "N/A",
                    "Start Time": pd.to_datetime(problem.get("startTime"), unit='ms'),
                    "End Time": pd.to_datetime(problem.get("endTime"), unit='ms') if problem.get("endTime") != -1 else "Ongoing",
                    "Management Zones": ", ".join(zone.get("name", "N/A") for zone in problem.get("managementZones", []))
                }
                problems.append(problem_info)

            df = pd.DataFrame(problems)

            
            output_file = "dynatrace_problems.xlsx"
            df.to_excel(output_file, index=False, sheet_name="Raw Data")

            
            wb = load_workbook(output_file)
            ws_raw = wb["Raw Data"]

            
            for cell in ws_raw[1]:  
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            
            ws_raw.auto_filter.ref = ws_raw.dimensions

            
            ws_pivot = wb.create_sheet(title="Repetetive Index")

            
            pivot_data = df.groupby(["Impact Level", "Severity Level"]).size().reset_index(name="Count")
            for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_pivot.cell(row=r_idx, column=c_idx, value=value)

            
            tab = Table(displayName="PivotTable1", ref=f"A1:C{len(pivot_data) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws_pivot.add_table(tab)

            
            wb.save(output_file)
            return send_file(output_file, as_attachment=True)

        else:
            return f"Failed to fetch data. Status code: {response.status_code}, Response: {response.text}"

    except Exception as e:
        return f"An error occurred: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)



# from flask import Flask, render_template, request, send_file
# import pandas as pd
# import requests
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import Font, PatternFill
# from openpyxl.worksheet.table import Table, TableStyleInfo
# import os

# app = Flask(__name__)

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/fetch', methods=['POST'])
# def fetch_data():
#     try:
#         # Get input from the user
#         url = request.form['url']
#         from_time = request.form['from']
#         to_time = request.form['to']
#         management_zone = request.form['managementZone']
#         token = request.form['token']

#         # Define API parameters and headers
#         params = {
#             "from": from_time,
#             "to": to_time,
#             "problemSelector": f'managementZones("{management_zone}")'
#         }
#         headers = {
#             "Authorization": f"Api-Token {token}"
#         }

#         # Fetch data from the API
#         response = requests.get(url, params=params, headers=headers)

#         if response.status_code == 200:
#             data = response.json()
#             problems = []

#             # Process API response
#             for problem in data.get("problems", []):
#                 problem_info = {
#                     "Problem ID": problem.get("problemId"),
#                     "Display ID": problem.get("displayId"),
#                     "Title": problem.get("title"),
#                     "Impact Level": problem.get("impactLevel"),
#                     "Severity Level": problem.get("severityLevel"),
#                     "Status": problem.get("status"),
#                     "Root Cause Entity": problem.get("rootCauseEntity", {}).get("name") if problem.get("rootCauseEntity") else "N/A",
#                     "Start Time": pd.to_datetime(problem.get("startTime"), unit='ms'),
#                     "End Time": pd.to_datetime(problem.get("endTime"), unit='ms') if problem.get("endTime") != -1 else "Ongoing",
#                     "Management Zones": ", ".join(zone.get("name", "N/A") for zone in problem.get("managementZones", []))
#                 }
#                 problems.append(problem_info)

#             df = pd.DataFrame(problems)

#             # Save the DataFrame to an Excel file
#             output_file = "dynatrace_problems.xlsx"
#             df.to_excel(output_file, index=False, sheet_name="Raw Data")

#             # Add pivot tables to the next sheet
#             wb = load_workbook(output_file)
#             ws_raw = wb["Raw Data"]

#             # Style the Raw Data sheet
#             for cell in ws_raw[1]:  # Apply styles to header row
#                 cell.font = Font(bold=True)
#                 cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#             # Create autofilter for the header row
#             ws_raw.auto_filter.ref = ws_raw.dimensions

#             # Create a new sheet for pivot tables
#             ws_pivot = wb.create_sheet(title="Repetetive Index")

#             # Add a simple summary pivot table
#             pivot_data = df.groupby(["Impact Level", "Severity Level"]).size().reset_index(name="Count")
#             for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=False, header=True), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     ws_pivot.cell(row=r_idx, column=c_idx, value=value)

#             # Style the pivot table
#             tab = Table(displayName="PivotTable1", ref=f"A1:C{len(pivot_data) + 1}")
#             style = TableStyleInfo(
#                 name="TableStyleMedium9",
#                 showFirstColumn=False,
#                 showLastColumn=False,
#                 showRowStripes=True,
#                 showColumnStripes=True,
#             )
#             tab.tableStyleInfo = style
#             ws_pivot.add_table(tab)

#             # Save the workbook
#             wb.save(output_file)
#             return send_file(output_file, as_attachment=True)

#         else:
#             return f"Failed to fetch data. Status code: {response.status_code}, Response: {response.text}"

#     except Exception as e:
#         return f"An error occurred: {str(e)}"

# if __name__ == '__main__':
#     app.run(host='0.0.0.0',debug=True)



# import requests
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.workbook import Workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.styles import Font, PatternFill

# # API details
# url = "https://guu84124.live.dynatrace.com/api/v2/problems"
# params = {
#     "from": "-7d",
#     "to": "now",
#     "problemSelector": "managementZones(\"EasyTravel\")"
# }
# headers = {
#     "Authorization": "Api-Token dt0c01.PODOCF76JULWPSP3ZFRTODYQ.XNTJ3YU4ZPOA7GYDVKFWHFKTDEFXLWMJFA3PSNVJ27YE7PM7CBIX4TPJCEO2KTPF"
# }

# # Fetch data from API
# response = requests.get(url, params=params, headers=headers)

# if response.status_code == 200:
#     data = response.json()

#     # Process the data into a format suitable for an Excel file
#     problems = []
#     for problem in data.get("problems", []):
#         problem_info = {
#             "Problem ID": problem.get("problemId"),
#             "Display ID": problem.get("displayId"),
#             "Title": problem.get("title"),
#             "Impact Level": problem.get("impactLevel"),
#             "Severity Level": problem.get("severityLevel"),
#             "Status": problem.get("status"),
#             "Root Cause Entity": problem.get("rootCauseEntity", {}).get("name") if problem.get("rootCauseEntity") else "N/A",
#             "Start Time": pd.to_datetime(problem.get("startTime"), unit='ms'),
#             "End Time": pd.to_datetime(problem.get("endTime"), unit='ms') if problem.get("endTime") != -1 else "Ongoing",
#             "Management Zones": ", ".join(zone.get("name", "N/A") for zone in problem.get("managementZones", []))
#         }
#         problems.append(problem_info)

#     # Convert to a Pandas DataFrame
#     df = pd.DataFrame(problems)

#     # Save the DataFrame to an Excel file
#     output_file = "dynatrace_problems.xlsx"
#     df.to_excel(output_file, index=False, sheet_name="Raw Data")

#     # Add pivot tables to the next sheet
#     wb = load_workbook(output_file)
#     ws_raw = wb["Raw Data"]

#     # Style the Raw Data sheet
#     for cell in ws_raw[1]:  # Apply styles to header row
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#     # Create autofilter for the header row
#     ws_raw.auto_filter.ref = ws_raw.dimensions

#     # Create a new sheet for pivot tables
#     ws_pivot = wb.create_sheet(title="Repetetive Index")

#     # Add a simple summary pivot table
#     pivot_data = df.groupby(["Impact Level", "Severity Level"]).size().reset_index(name="Count")
#     for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=False, header=True), 1):
#         for c_idx, value in enumerate(row, 1):
#             ws_pivot.cell(row=r_idx, column=c_idx, value=value)

#     # Style the pivot table
#     tab = Table(displayName="PivotTable1", ref=f"A1:C{len(pivot_data) + 1}")
#     style = TableStyleInfo(
#         name="TableStyleMedium9",
#         showFirstColumn=False,
#         showLastColumn=False,
#         showRowStripes=True,
#         showColumnStripes=True,
#     )
#     tab.tableStyleInfo = style
#     ws_pivot.add_table(tab)

#     # Save the workbook
#     wb.save(output_file)
#     print(f"Data and pivot tables have been successfully saved to {output_file}")
# else:
#     print(f"Failed to fetch data. Status code: {response.status_code}, Response: {response.text}")
